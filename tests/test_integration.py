"""Integration tests for ExcelCppSyntaxHighlight."""

import pytest
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from cpp_highlight import (
    is_cpp_code,
    CellHighlighter,
    ThemeConfig,
    FontSettings,
)
from cpp_highlight.processor import process_excel


class TestIntegration:
    """End-to-end integration tests."""

    @pytest.fixture
    def sample_workbook(self, tmp_path):
        """Create a sample workbook with C++ code."""
        wb = Workbook()
        ws = wb.active

        # Cell with C++ code
        ws["A1"] = """#include <iostream>
int main() {
    std::cout << "Hello" << std::endl;
    return 0;
}"""

        # Cell with plain text
        ws["A2"] = "This is just plain text"

        # Cell with another C++ snippet
        ws["A3"] = "std::vector<int> numbers = {1, 2, 3};"

        path = tmp_path / "input.xlsx"
        wb.save(path)
        return path

    def test_process_excel_highlights_cpp_cells(self, sample_workbook, tmp_path):
        """Processing Excel file highlights C++ cells."""
        output_path = tmp_path / "output.xlsx"

        count = process_excel(str(sample_workbook), str(output_path), verbose=False)

        assert count == 2  # A1 and A3 are C++ code

        # Verify output file exists and is valid
        assert output_path.exists()
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Note: CellRichText is serialized as plain text when saved to xlsx,
        # so we verify the content and row heights instead

        # A1 should contain the original code
        assert "#include" in str(ws["A1"].value)
        assert "std::cout" in str(ws["A1"].value)

        # A1 row height should be adjusted (multiline code)
        assert ws.row_dimensions[1].height is not None
        assert ws.row_dimensions[1].height > 16  # More than single line

        # A2 should still be plain text (unchanged)
        assert ws["A2"].value == "This is just plain text"

        # A3 row height should be adjusted (single line, minimal height)
        assert ws.row_dimensions[3].height is not None

    def test_cell_highlighter_with_custom_theme(self):
        """CellHighlighter works with custom theme."""
        custom_colors = ThemeConfig(colors={"Keyword": "FF0000"})
        highlighter = CellHighlighter(theme=custom_colors)

        rich_text, height = highlighter.highlight("int x = 42;")

        assert rich_text is not None
        assert height is not None
        assert height > 0

    def test_cell_highlighter_with_custom_font(self):
        """CellHighlighter works with custom font settings."""
        custom_font = FontSettings(name="Courier New", size=14)
        highlighter = CellHighlighter(font=custom_font)

        rich_text, height = highlighter.highlight("int x = 42;")

        assert rich_text is not None

    def test_empty_workbook(self, tmp_path):
        """Processing empty workbook produces valid output."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = ""

        input_path = tmp_path / "empty.xlsx"
        output_path = tmp_path / "output.xlsx"
        wb.save(input_path)

        count = process_excel(str(input_path), str(output_path))

        assert count == 0
        assert output_path.exists()

    def test_multiline_code_height_calculation(self):
        """Multiline code gets correct height."""
        highlighter = CellHighlighter()

        code = "line1\nline2\nline3\nline4\nline5"
        _, height = highlighter.highlight(code)

        # 5 lines = base_height + 4 * line_height
        expected = 16.0 + 4 * 16.0
        assert height == expected

    def test_cli_entry_point(self, sample_workbook, tmp_path):
        """CLI entry point works correctly."""
        import subprocess
        import sys

        output_path = tmp_path / "output.xlsx"

        result = subprocess.run(
            [
                sys.executable,
                "-m",
                "cpp_highlight",
                str(sample_workbook),
                "-o",
                str(output_path),
            ],
            capture_output=True,
            text=True,
            cwd="/mnt/d/GitRepos/ExcelCppSyntaxHighlight",
        )

        assert result.returncode == 0
        assert "Processed" in result.stdout
        assert output_path.exists()
