# cpp_highlight/processor.py
"""Excel file processing logic."""

import sys

import openpyxl

from cpp_highlight.config import FontSettings, ThemeConfig
from cpp_highlight.core import CellHighlighter, is_cpp_code
from cpp_highlight.core.highlighter import calculate_required_height


def process_excel(input_path: str, output_path: str, verbose: bool = False) -> int:
    """Process an Excel file and apply C++ syntax highlighting.

    Args:
        input_path: Path to input Excel file
        output_path: Path to output Excel file
        verbose: Enable verbose output

    Returns:
        Number of cells highlighted
    """
    if verbose:
        print(f"Loading: {input_path}")

    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(f"Error: Failed to load workbook: {e}", file=sys.stderr)
        sys.exit(1)

    highlighter = CellHighlighter()
    highlighted_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if verbose:
            print(f"\nProcessing sheet: {sheet_name}")

        row_height_requirements = {}

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and is_cpp_code(cell.value):
                    if verbose:
                        print(f"  {cell.coordinate}: Detected C++ code")

                    rich_text, required_height = highlighter.highlight(cell.value)
                    if rich_text is not None:
                        cell.value = rich_text
                        from openpyxl.styles import Alignment

                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        highlighted_count += 1
                        if verbose:
                            print(f"    -> Highlighted")

                        if required_height is not None:
                            row_num = cell.row
                            current_max = row_height_requirements.get(row_num, 0)
                            row_height_requirements[row_num] = max(
                                current_max, required_height
                            )

        for row_num, required_height in row_height_requirements.items():
            original_height = ws.row_dimensions[row_num].height

            if original_height is None:
                ws.row_dimensions[row_num].height = required_height
            else:
                ws.row_dimensions[row_num].height = max(
                    original_height, required_height
                )

    if verbose:
        print(f"\nSaving: {output_path}")

    try:
        wb.save(output_path)
    except Exception as e:
        print(f"Error: Failed to save workbook: {e}", file=sys.stderr)
        sys.exit(1)

    return highlighted_count
