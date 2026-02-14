#!/usr/bin/env python3
"""
C++ Excel Highlighter

A command-line tool that automatically detects C++ code in Excel cells
and applies syntax highlighting using the Atom One Light color theme.
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Color, Alignment
from pygments import lex
from pygments.lexers import CppLexer
from pygments.token import Token
from xml.etree.ElementTree import Element

import json

from .config import FontSettings
from .core.detection import is_cpp_code, C_DETECTORS_HIGH, C_DETECTORS_MEDIUM


# Custom TextBlock that always sets xml:space="preserve" for whitespace tokens
# We must inherit from the original TextBlock for CellRichText type checking
from openpyxl.cell.rich_text import TextBlock as _OrigTextBlock


class TextBlock(_OrigTextBlock):
    """TextBlock that properly preserves whitespace in Excel."""

    def to_tree(self):
        """Convert to XML tree, ensuring whitespace is preserved."""
        el = Element("r")
        if self.font:
            el.append(self.font.to_tree(tagname="rPr"))
        t = Element("t")
        t.text = self.text

        # Always set xml:space="preserve" if text contains any whitespace
        if self.text and (self.text != self.text.strip() or not self.text.strip()):
            XML_NS = "http://www.w3.org/XML/1998/namespace"
            t.set("{%s}space" % XML_NS, "preserve")

        el.append(t)
        return el


# Token type to string mapping for JSON config lookup
TOKEN_TYPE_NAMES = {
    Token.Keyword: "Keyword",
    Token.Keyword.Type: "Keyword.Type",
    Token.Keyword.Namespace: "Keyword.Namespace",
    Token.Keyword.Constant: "Keyword.Constant",
    Token.Keyword.Declaration: "Keyword.Declaration",
    Token.String: "String",
    Token.String.Single: "String.Single",
    Token.String.Double: "String.Double",
    Token.String.Doc: "String.Doc",
    Token.Comment: "Comment",
    Token.Comment.Single: "Comment.Single",
    Token.Comment.Multiline: "Comment.Multiline",
    Token.Comment.Preproc: "Comment.Preproc",
    Token.Comment.PreprocFile: "Comment.PreprocFile",
    Token.Number: "Number",
    Token.Number.Integer: "Number.Integer",
    Token.Number.Float: "Number.Float",
    Token.Number.Hex: "Number.Hex",
    Token.Number.Bin: "Number.Bin",
    Token.Number.Oct: "Number.Oct",
    Token.Name.Function: "Name.Function",
    Token.Name.Class: "Name.Class",
    Token.Name.Namespace: "Name.Namespace",
    Token.Name.Builtin: "Name.Builtin",
    Token.Name.Builtin.Pseudo: "Name.Builtin.Pseudo",
    Token.Name.Decorator: "Name.Decorator",
    Token.Name: "Name",
    Token.Name.Attribute: "Name.Attribute",
    Token.Name.Tag: "Name.Tag",
    Token.Name.Variable: "Name.Variable",
    Token.Name.Variable.Class: "Name.Variable.Class",
    Token.Name.Variable.Global: "Name.Variable.Global",
    Token.Name.Variable.Instance: "Name.Variable.Instance",
    Token.Name.Variable.Magic: "Name.Variable.Magic",
    Token.Name.Other: "Name.Other",
    Token.Operator: "Operator",
    Token.Operator.Word: "Operator.Word",
    Token.Punctuation: "Punctuation",
    Token.Text: "Text",
    Token.Text.Whitespace: "Text.Whitespace",
    Token.Error: "Error",
    Token.Other: "Other",
}


# Default theme configuration (fallback if JSON config fails)
DEFAULT_THEME_COLORS = {
    "Keyword": "A626A4",
    "Keyword.Type": "A626A4",
    "Keyword.Namespace": "A626A4",
    "Keyword.Constant": "A626A4",
    "Keyword.Declaration": "A626A4",
    "String": "50A14F",
    "String.Single": "50A14F",
    "String.Double": "50A14F",
    "String.Doc": "A0A1A7",
    "Comment": "A0A1A7",
    "Comment.Single": "A0A1A7",
    "Comment.Multiline": "A0A1A7",
    "Comment.Preproc": "A626A4",
    "Comment.PreprocFile": "A626A4",
    "Number": "986801",
    "Number.Integer": "986801",
    "Number.Float": "986801",
    "Number.Hex": "986801",
    "Number.Bin": "986801",
    "Number.Oct": "986801",
    "Name.Function": "4078F2",
    "Name.Class": "C18401",
    "Name.Namespace": "C18401",
    "Name.Builtin": "C18401",
    "Name.Builtin.Pseudo": "C18401",
    "Name.Decorator": "C18401",
    "Name": "E45649",
    "Name.Attribute": "E45649",
    "Name.Tag": "E45649",
    "Name.Variable": "E45649",
    "Name.Variable.Class": "E45649",
    "Name.Variable.Global": "E45649",
    "Name.Variable.Instance": "E45649",
    "Name.Variable.Magic": "E45649",
    "Name.Other": "E45649",
    "Operator": "383A42",
    "Operator.Word": "A626A4",
    "Punctuation": "383A42",
    "Text": "383A42",
    "Text.Whitespace": "383A42",
    "Error": "FF0000",
    "Other": "383A42",
}


def get_config_path():
    """Get the path to the theme configuration file."""
    if getattr(sys, "frozen", False):
        # Running in a bundled executable
        base_path = Path(sys.executable).parent
    else:
        # Running as a Python script
        base_path = Path(__file__).parent
    return base_path / "theme.json"


def create_default_config(config_path):
    """Create the default theme configuration file."""
    try:
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_THEME_COLORS, f, indent=2)
    except IOError as e:
        print(f"Warning: Failed to create default config: {e}", file=sys.stderr)


def load_theme_config():
    """Load theme configuration from JSON file.

    Returns a dictionary mapping token type names to hex color codes.
    If the config file doesn't exist, creates it with default values.
    If loading fails, returns the default theme.
    """
    config_path = get_config_path()

    # Create default config if it doesn't exist
    if not config_path.exists():
        create_default_config(config_path)

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"Warning: Failed to load theme config: {e}", file=sys.stderr)
        return DEFAULT_THEME_COLORS.copy()


# Load theme configuration at module import time
THEME_CONFIG = load_theme_config()


def get_color(token_type):
    """Get color for a token type from the theme configuration."""
    # Try exact match first
    type_name = TOKEN_TYPE_NAMES.get(token_type)
    if type_name and type_name in THEME_CONFIG:
        return THEME_CONFIG[type_name]

    # Try parent type matching
    for ttype, name in TOKEN_TYPE_NAMES.items():
        if token_type in ttype and name in THEME_CONFIG:
            return THEME_CONFIG[name]

    # Default color
    return "383A42"


def highlight_cell(cell):
    """Apply syntax highlighting to a cell containing C++ code.

    Returns:
        tuple: (success: bool, required_height: float or None)
        required_height is the calculated row height needed for this cell's content.
    """
    text = cell.value
    if not isinstance(text, str):
        return False, None

    if not is_cpp_code(text):
        return False, None

    try:
        tokens = list(lex(text, CppLexer()))

        # Remove trailing newline token added by CppLexer
        if tokens and tokens[-1][0] == Token.Text.Whitespace and tokens[-1][1] == "\n":
            tokens = tokens[:-1]

        blocks = []
        _font_settings = FontSettings.default()
        for token_type, value in tokens:
            color_hex = get_color(token_type)
            color_obj = Color(rgb=color_hex)
            font = InlineFont(
                color=color_obj, rFont=_font_settings.name, sz=_font_settings.size
            )
            block = TextBlock(text=value, font=font)
            blocks.append(block)

        rich_text = CellRichText(*blocks)
        cell.value = rich_text

        # Enable text wrapping to display multiline code correctly
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Calculate required row height for this cell
        required_height = calculate_required_height(tokens)

        return True, required_height

    except Exception as e:
        print(
            f"Warning: Failed to highlight cell {cell.coordinate}: {e}", file=sys.stderr
        )
        return False, None


def calculate_required_height(tokens):
    """Calculate the required row height for given tokens.

    Returns:
        float: Required row height in points.
    """
    # Count lines in the content
    line_count = 1
    for token_type, value in tokens:
        if "\n" in value:
            line_count += value.count("\n")

    # Calculate height based on line count
    # Excel default row height is 15 points for 11pt Calibri (1 line)
    # For 11pt Consolas code with proper readability, use 16 points per line
    # Base height: 16 points, each additional line adds 16 points
    _settings = FontSettings.default()
    base_height = _settings.base_height
    line_height = _settings.line_height

    return base_height + (line_count - 1) * line_height


def process_excel(input_path, output_path, verbose=False):
    """Process an Excel file and apply C++ syntax highlighting."""
    if verbose:
        print(f"Loading: {input_path}")

    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(f"Error: Failed to load workbook: {e}", file=sys.stderr)
        sys.exit(1)

    highlighted_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if verbose:
            print(f"\nProcessing sheet: {sheet_name}")

        # Track row height requirements: {row_number: max_required_height}
        row_height_requirements = {}

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and is_cpp_code(cell.value):
                    if verbose:
                        print(f"  {cell.coordinate}: Detected C++ code")

                    success, required_height = highlight_cell(cell)
                    if success:
                        highlighted_count += 1
                        if verbose:
                            print(f"    -> Highlighted")

                        # Track the maximum required height for this row
                        if required_height is not None:
                            row_num = cell.row
                            current_max = row_height_requirements.get(row_num, 0)
                            row_height_requirements[row_num] = max(
                                current_max, required_height
                            )

        # Apply row heights: compare original height with required height, take the larger
        for row_num, required_height in row_height_requirements.items():
            original_height = ws.row_dimensions[row_num].height

            if original_height is None:
                # No original height set, use required height
                ws.row_dimensions[row_num].height = required_height
            else:
                # Take the larger of original and required
                final_height = max(original_height, required_height)
                ws.row_dimensions[row_num].height = final_height

    if verbose:
        print(f"\nSaving: {output_path}")

    try:
        wb.save(output_path)
    except Exception as e:
        print(f"Error: Failed to save workbook: {e}", file=sys.stderr)
        sys.exit(1)

    return highlighted_count


def main():
    parser = argparse.ArgumentParser(
        description="Apply C++ syntax highlighting to Excel cells",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python cpp_highlight.py input.xlsx -o output.xlsx
  python cpp_highlight.py code.xlsx --output highlighted.xlsx --verbose
        """,
    )

    parser.add_argument("input", help="Input Excel file path")
    parser.add_argument("-o", "--output", required=True, help="Output Excel file path")
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="Enable verbose output"
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    if not input_path.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        print(
            f"Warning: Input file may not be a valid Excel file: {args.input}",
            file=sys.stderr,
        )

    count = process_excel(args.input, args.output, args.verbose)

    print(f"Processed {count} cells with C++ code")
    print(f"  Input:  {args.input}")
    print(f"  Output: {args.output}")


if __name__ == "__main__":
    main()
