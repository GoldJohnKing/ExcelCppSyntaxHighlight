#!/usr/bin/env python3
"""Create test Excel file with C++ code samples."""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill


def create_test_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Add test cases
    test_cases = [
        (
            "Hello World",
            """#include <iostream>

int main() {
    std::cout << "Hello, World!" << std::endl;
    return 0;
}""",
        ),
        (
            "Class Example",
            """class Person {
private:
    std::string name;
    int age;
public:
    Person(std::string n, int a) : name(n), age(a) {}
    void greet() {
        std::cout << "Hello, I'm " << name << std::endl;
    }
};""",
        ),
        (
            "Template",
            """template<typename T>
T max(T a, T b) {
    return (a > b) ? a : b;
}

int main() {
    int x = max(3, 5);
    double y = max(2.5, 1.5);
    return 0;
}""",
        ),
        (
            "Comments",
            """// This is a single line comment
int x = 10; /* inline comment */
/* Multi-line
   comment block */
const int MAX_SIZE = 100;""",
        ),
        (
            "Control Flow",
            """for (int i = 0; i < 10; i++) {
    if (i % 2 == 0) {
        continue;
    }
    while (x > 0) {
        x--;
    }
}
switch (value) {
    case 1: break;
    default: return;
}""",
        ),
        (
            "Lambda",
            """auto lambda = [](int x, int y) -> int {
    return x + y;
};

std::vector<int> nums = {1, 2, 3, 4, 5};
std::sort(nums.begin(), nums.end(), [](int a, int b) {
    return a > b;
});""",
        ),
        ("Not Code", "This is just a regular text cell with no code."),
        ("Mixed", "Here is some code: int x = 42; and more text after."),
        (
            "Syntax Error",
            """int main() {
    int x =     // Missing value
    if x > 0 {  // Missing parentheses
        return
    }           // Missing semicolon
}""",
        ),
        ("Unclosed String", 'string s = "hello'),  # Intentionally unclosed
        (
            "STL Containers",
            """std::vector<int> vec = {1, 2, 3};
std::map<std::string, int> scores;
std::set<double> values;
vec.push_back(4);
scores["Alice"] = 95;""",
        ),
    ]

    # Define styles for testing format preservation
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for idx, (title, code) in enumerate(test_cases, 1):
        cell_title = ws.cell(row=idx, column=1, value=title)
        cell_code = ws.cell(row=idx, column=2, value=code)

        # Set some formatting to test preservation
        if idx % 2 == 0:
            cell_title.alignment = Alignment(horizontal="center")
            cell_code.alignment = Alignment(horizontal="left", vertical="center")

        # Add borders to some cells
        if idx % 3 == 0:
            cell_code.border = thin_border

        # Add fill color to some cells
        if idx == 1:
            cell_title.fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60

    wb.save("test_input.xlsx")
    print("Created test_input.xlsx with 11 test cases")


if __name__ == "__main__":
    create_test_file()
