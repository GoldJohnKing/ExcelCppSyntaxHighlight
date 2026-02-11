#!/usr/bin/env python3
"""
C++ Excel Highlighter

A command-line tool that automatically detects C++ code in Excel cells
and applies syntax highlighting using the Atom One Light color theme.
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Color, Alignment
from pygments import lex
from pygments.lexers import CppLexer
from pygments.token import Token
from xml.etree.ElementTree import Element


# Custom TextBlock that always sets xml:space="preserve" for whitespace tokens
# We must inherit from the original TextBlock for CellRichText type checking
from openpyxl.cell.rich_text import TextBlock as _OrigTextBlock


class TextBlock(_OrigTextBlock):
    """TextBlock that properly preserves whitespace in Excel."""

    def to_tree(self):
        """Convert to XML tree, ensuring whitespace is preserved."""
        el = Element("r")
        if self.font:
            el.append(self.font.to_tree(tagname="rPr"))
        t = Element("t")
        t.text = self.text

        # Always set xml:space="preserve" if text contains any whitespace
        if self.text and (self.text != self.text.strip() or not self.text.strip()):
            XML_NS = "http://www.w3.org/XML/1998/namespace"
            t.set("{%s}space" % XML_NS, "preserve")

        el.append(t)
        return el


# Atom One Light color theme
ATOM_ONE_LIGHT_THEME = {
    # Keywords - Purple
    Token.Keyword: "A626A4",
    Token.Keyword.Type: "A626A4",
    Token.Keyword.Namespace: "A626A4",
    Token.Keyword.Constant: "A626A4",
    Token.Keyword.Declaration: "A626A4",
    # Strings - Green
    Token.String: "50A14F",
    Token.String.Single: "50A14F",
    Token.String.Double: "50A14F",
    Token.String.Doc: "A0A1A7",
    # Comments - Gray
    Token.Comment: "A0A1A7",
    Token.Comment.Single: "A0A1A7",
    Token.Comment.Multiline: "A0A1A7",
    Token.Comment.Preproc: "A626A4",
    Token.Comment.PreprocFile: "A626A4",
    # Numbers - Gold
    Token.Number: "986801",
    Token.Number.Integer: "986801",
    Token.Number.Float: "986801",
    Token.Number.Hex: "986801",
    Token.Number.Bin: "986801",
    Token.Number.Oct: "986801",
    # Functions - Blue
    Token.Name.Function: "4078F2",
    # Classes and Types - Gold
    Token.Name.Class: "C18401",
    Token.Name.Namespace: "C18401",
    Token.Name.Builtin: "C18401",
    Token.Name.Builtin.Pseudo: "C18401",
    Token.Name.Decorator: "C18401",
    # Variables - Red
    Token.Name: "E45649",
    Token.Name.Attribute: "E45649",
    Token.Name.Tag: "E45649",
    Token.Name.Variable: "E45649",
    Token.Name.Variable.Class: "E45649",
    Token.Name.Variable.Global: "E45649",
    Token.Name.Variable.Instance: "E45649",
    Token.Name.Variable.Magic: "E45649",
    Token.Name.Other: "E45649",
    # Operators and punctuation - Dark Gray
    Token.Operator: "383A42",
    Token.Operator.Word: "A626A4",
    Token.Punctuation: "383A42",
    # Generic text - Dark Gray
    Token.Text: "383A42",
    Token.Text.Whitespace: "383A42",
    Token.Error: "FF0000",
    Token.Other: "383A42",
}


# C++ detection patterns
C_DETECTORS_HIGH = [
    r'#include\s*[\<"]',
    r"using\s+namespace\s+\w+",
    r"int\s+main\s*\(",
    r"std::",
    r"::\s*\w+\s*\(",
    r"template\s*\<",
]

C_DETECTORS_MEDIUM = [
    r"\b(class|struct|enum)\s+\w+",
    r"\b(int|char|float|double|void|bool|auto|const|constexpr|mutable)\b",
    r"^\s*(public|private|protected):",
    r"^\s*#\s*(define|ifdef|ifndef|endif|pragma)",
    r"\b(for|while|if|else|switch|case|break|continue|return)\b",
    r"\b(cout|cin|endl|printf|scanf)\b",
    r"\<\<|\>\>",
    r"\b(string|vector|map|set|array)\b",
    r"//",  # Single line comment
    r"/\*.*?\*/",  # Multi-line comment (non-greedy)
]


def get_color(token_type):
    """Get color for a token type from the Atom One Light theme."""
    if token_type in ATOM_ONE_LIGHT_THEME:
        return ATOM_ONE_LIGHT_THEME[token_type]

    for ttype, color in ATOM_ONE_LIGHT_THEME.items():
        if token_type in ttype:
            return color

    return "383A42"


def is_cpp_code(text):
    """Detect if text contains C++ code."""
    if not text or not isinstance(text, str):
        return False

    high_confidence = 0
    medium_confidence = 0

    for pattern in C_DETECTORS_HIGH:
        matches = re.findall(pattern, text, re.MULTILINE)
        high_confidence += len(matches)

    for pattern in C_DETECTORS_MEDIUM:
        flags = re.MULTILINE
        if "/*" in pattern:
            flags |= re.DOTALL
        matches = re.findall(pattern, text, flags)
        medium_confidence += len(matches)

    if high_confidence >= 1:
        return True
    if medium_confidence >= 3:
        return True
    if medium_confidence >= 2 and len(text) > 100:
        return True

    return False


def highlight_cell(cell):
    """Apply syntax highlighting to a cell containing C++ code."""
    text = cell.value
    if not isinstance(text, str):
        return False

    if not is_cpp_code(text):
        return False

    try:
        tokens = list(lex(text, CppLexer()))

        blocks = []
        for token_type, value in tokens:
            color_hex = get_color(token_type)
            color_obj = Color(rgb=color_hex)
            font = InlineFont(color=color_obj, rFont="Consolas", sz=11)
            block = TextBlock(text=value, font=font)
            blocks.append(block)

        rich_text = CellRichText(*blocks)
        cell.value = rich_text

        # Enable text wrapping to display multiline code correctly
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        return True

    except Exception as e:
        print(
            f"Warning: Failed to highlight cell {cell.coordinate}: {e}", file=sys.stderr
        )
        return False


def process_excel(input_path, output_path, verbose=False):
    """Process an Excel file and apply C++ syntax highlighting."""
    if verbose:
        print(f"Loading: {input_path}")

    try:
        wb = openpyxl.load_workbook(input_path)
    except Exception as e:
        print(f"Error: Failed to load workbook: {e}", file=sys.stderr)
        sys.exit(1)

    highlighted_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if verbose:
            print(f"\nProcessing sheet: {sheet_name}")

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and is_cpp_code(cell.value):
                    if verbose:
                        print(f"  {cell.coordinate}: Detected C++ code")

                    if highlight_cell(cell):
                        highlighted_count += 1
                        if verbose:
                            print(f"    ✓ Highlighted")

    if verbose:
        print(f"\nSaving: {output_path}")

    try:
        wb.save(output_path)
    except Exception as e:
        print(f"Error: Failed to save workbook: {e}", file=sys.stderr)
        sys.exit(1)

    return highlighted_count


def main():
    parser = argparse.ArgumentParser(
        description="Apply C++ syntax highlighting to Excel cells",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python cpp_highlight.py input.xlsx -o output.xlsx
  python cpp_highlight.py code.xlsx --output highlighted.xlsx --verbose
        """,
    )

    parser.add_argument("input", help="Input Excel file path")
    parser.add_argument("-o", "--output", required=True, help="Output Excel file path")
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="Enable verbose output"
    )

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    if not input_path.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        print(
            f"Warning: Input file may not be a valid Excel file: {args.input}",
            file=sys.stderr,
        )

    count = process_excel(args.input, args.output, args.verbose)

    print(f"✓ Processed {count} cells with C++ code")
    print(f"  Input:  {args.input}")
    print(f"  Output: {args.output}")


if __name__ == "__main__":
    main()
